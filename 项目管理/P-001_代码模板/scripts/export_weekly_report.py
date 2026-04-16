# -*- coding: utf-8 -*-
"""
生成并导出周报为Excel文件
包含用户数据、反馈统计、排行榜等
"""

import os
import sys
import logging
from pathlib import Path
from datetime import datetime, timedelta

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 添加项目路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from modules.sqlite_manager import SQLiteManager


def export_to_excel(filename=None):
    """
    导出周报为Excel文件

    Args:
        filename: 导出文件名（可选，默认为带时间戳）
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("❌ 需要安装 openpyxl: pip install openpyxl")
        return False

    logger.info("=" * 60)
    logger.info("📊 开始生成周报Excel")
    logger.info("=" * 60)

    db = SQLiteManager()

    try:
        # 创建workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 移除默认sheet

        # 1. 概览统计 Sheet
        logger.info("📝 生成概览统计...")
        create_overview_sheet(wb, db)

        # 2. 用户列表 Sheet
        logger.info("📝 生成用户列表...")
        create_users_sheet(wb, db)

        # 3. 反馈汇总 Sheet
        logger.info("📝 生成反馈汇总...")
        create_feedback_sheet(wb, db)

        # 4. 排行榜 Sheet
        logger.info("📝 生成排行榜...")
        create_ranking_sheet(wb, db)

        # 保存文件
        export_dir = Path('reports')
        export_dir.mkdir(exist_ok=True)

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"weekly_report_{timestamp}.xlsx"

        filepath = export_dir / filename
        wb.save(str(filepath))

        logger.info(f"✅ 周报已导出: {filepath}")
        logger.info(f"📌 文件大小: {os.path.getsize(filepath)} 字节")

        return True

    except Exception as e:
        logger.error(f"❌ 导出失败: {e}", exc_info=True)
        return False

    finally:
        db.close()


def create_overview_sheet(wb, db):
    """创建概览统计Sheet"""
    ws = wb.create_sheet('📊 概览', 0)

    # 标题
    ws['A1'] = "邀测招募 - 周报"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    # 生成日期
    ws['A2'] = f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')

    # 统计数据
    stats = db.get_stats()

    data = [
        ['指标', '数值', '', ''],
        ['总用户数', stats.get('total_users', 0), '', ''],
        ['本周新增', stats.get('weekly_new_users', 0), '', ''],
        ['总反馈数', stats.get('total_feedbacks', 0), '', ''],
        ['本周新反馈', stats.get('weekly_new_feedbacks', 0), '', ''],
        ['总邀请数', stats.get('total_invitations', 0), '', ''],
        ['本周新邀请', stats.get('weekly_new_invitations', 0), '', ''],
        ['转化率', f"{stats.get('conversion_rate', 0):.1%}", '', ''],
    ]

    for row_idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 4:  # 表头
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')

    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12


def create_users_sheet(wb, db):
    """创建用户列表Sheet"""
    ws = wb.create_sheet('👥 用户', 1)

    # 表头
    headers = ['序号', '用户名', '来源', '邮箱/手机', '加入日期', '积分', '状态']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    # 用户数据
    users = db.get_all_users()
    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=user.get('昵称', ''))
        ws.cell(row=row_idx, column=3, value=user.get('来源渠道', ''))
        ws.cell(row=row_idx, column=4, value=user.get('联系方式', ''))
        ws.cell(row=row_idx, column=5, value=user.get('创建时间', '')[:10])
        ws.cell(row=row_idx, column=6, value=user.get('积分', 0))
        ws.cell(row=row_idx, column=7, value='正常')

    # 调整列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8


def create_feedback_sheet(wb, db):
    """创建反馈汇总Sheet"""
    ws = wb.create_sheet('💬 反馈', 2)

    # 表头
    headers = ['序号', '用户', '反馈类型', '反馈内容', '日期', '状态']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    # 反馈数据
    feedbacks = db.get_all_feedback()
    for row_idx, fb in enumerate(feedbacks, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=fb.get('用户名', ''))
        ws.cell(row=row_idx, column=3, value=fb.get('分类', 'other'))
        ws.cell(row=row_idx, column=4, value=fb.get('反馈内容', '')[:50])  # 截断显示
        ws.cell(row=row_idx, column=5, value=fb.get('创建时间', '')[:10])
        ws.cell(row=row_idx, column=6, value='已读')

    # 调整列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8


def create_ranking_sheet(wb, db):
    """创建排行榜Sheet"""
    ws = wb.create_sheet('🏆 排行', 3)

    # 表头
    headers = ['排名', '用户', '邀请数', '反馈数', '总积分']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')

    # 获取排行数据
    users = db.get_all_users()
    sorted_users = sorted(users, key=lambda x: x.get('积分', 0), reverse=True)

    medals = ['🥇', '🥈', '🥉']

    for rank, user in enumerate(sorted_users[:50], 1):  # 显示top 50
        medal = medals[rank - 1] if rank <= 3 else f"{rank}."
        ws.cell(row=rank + 1, column=1, value=f"{medal} {rank}")
        ws.cell(row=rank + 1, column=2, value=user.get('昵称', ''))
        ws.cell(row=rank + 1, column=3, value=user.get('邀请数', 0))
        ws.cell(row=rank + 1, column=4, value=user.get('反馈数', 0))
        ws.cell(row=rank + 1, column=5, value=user.get('积分', 0))

        # Top 3 高亮
        if rank <= 3:
            for col in range(1, 6):
                ws.cell(row=rank + 1, column=col).fill = PatternFill(
                    start_color='FFFF99', end_color='FFFF99', fill_type='solid'
                )

    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10


if __name__ == '__main__':
    success = export_to_excel()
    sys.exit(0 if success else 1)
